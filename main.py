import sqlite3

import openpyxl
from flask import Flask, render_template, redirect, make_response, request, session, abort, jsonify
from data.users import User
from forms.user import RegisterForm, LoginForm
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from data import db_session
from openpyxl import Workbook
from datetime import datetime
#
app = Flask(__name__)
login_manager = LoginManager()
login_manager.init_app(app)
app.config['SECRET_KEY'] = 'yandexlyceum_secret_key'
conn1 = sqlite3.connect('db/teachers.db', check_same_thread=False)
cursor1 = conn1.cursor()
teachers = cursor1.execute('SELECT name, school FROM teacher').fetchall()
e = 0


def get_day_of_week(date):
    # Задаем дату
    date_str = date
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()

    # Получаем номер дня недели (0 - понедельник, 1 - вторник и т.д.)
    weekday_num = date_obj.weekday()

    # Массив с названиями дней недели
    weekdays = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

    # Выводим название дня недели
    return weekdays[weekday_num]


def m():
    global lessons, clas, urok, students, selected_value, r, wb, ws, online_user
    conn = sqlite3.connect('db/users.db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS School (
            id INTEGER PRIMARY KEY,
            school TEXT UNIQUE
        )
    ''')
    conn.commit()
    conn.close()
    students = {}
    conn = sqlite3.connect('db/schedule.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM lessons WHERE school=? AND ticher LIKE ?', (online_user.school, online_user.name))
    lessons = cursor.fetchall()
    conn.commit()
    conn.close()
    urok = []
    clas = []
    for el in lessons:
        if el[-1] not in clas:
            clas.append(el[-1])
        if el[3] not in urok:
            urok.append(el[3])
    selected_value = 'Select....'
    r = selected_value
    wb = Workbook()
    ws = wb.active
    ws.append(['Student', 'Subject', 'Grade', 'Date to', 'Date From'])


@app.errorhandler(404)
def not_found(error):
    return make_response(jsonify({'error': 'Not found'}), 404)


@login_manager.user_loader
def load_user(user_id):
    db_sess = db_session.create_session()
    return db_sess.query(User).get(user_id)


@app.route('/grades')
def grades():
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = {}

    for row in sheet.iter_rows(values_only=True):
        name, subject, grade, dat = row[:4]  # Предполагается, что дата и оценка идут первыми двумя столбцами
        dat = str(dat).split()[0]
        if name == online_user.name:
            if subject not in data:
                data[dat] = []
            data[dat].append({subject: grade})
    data = dict(sorted(data.items(), reverse=True))
    return render_template('grade.html', grades=data)


@app.route('/register', methods=['GET', 'POST'])
def registration():
    form = RegisterForm()
    if form.validate_on_submit():
        if 1 > form.num_class.data or form.num_class.data > 11:
            return render_template('register.html', title='Регистрация',
                                   form=form,
                                   message="Не правильная цифра класса")

        if 1 > form.school.data:
            return render_template('register.html', title='Регистрация',
                                   form=form,
                                   message="Не правильный номер школы")

        if len(form.alfa_class.data) > 1 or len(form.alfa_class.data) == 0:
            return render_template('register.html', title='Регистрация',
                                   form=form,
                                   message="Не правильная буква класса")

        if form.password.data != form.password_again.data:
            return render_template('register.html', title='Регистрация',
                                   form=form,
                                   message="Пароли не совпадают")

        if len(form.first_name.data) == 0 or len(form.second_name.data) == 0:
            return render_template('register.html', title='Регистрация',
                                   form=form,
                                   message="Не Правильные Имя или Фамилия")
        if form.types.data == 'Учитель' and (
                f'{form.second_name.data.capitalize()} {form.first_name.data.capitalize()}',
                form.school.data) not in teachers:
            return render_template('register.html', title='Регистрация',
                                   form=form,
                                   message="Нет такого учителя")

        db_sess = db_session.create_session()

        if db_sess.query(User).filter(User.email == form.email.data).first():
            return render_template('register.html', title='Регистрация',
                                   form=form,
                                   message="Такой пользователь уже есть")

        user = User(
            name=form.first_name.data + ' ' + form.second_name.data,
            email=form.email.data,
            types=form.types.data,
            classes=f'{form.num_class.data}{form.alfa_class.data.upper()}',
            school=form.school.data
        )
        user.set_password(form.password.data)
        db_sess.add(user)
        db_sess.commit()
        conn = sqlite3.connect('db/users.db')
        cursor = conn.cursor()
        cursor.execute(f'INSERT School(school) VALUES({user.classes})')
        return redirect('/login')
    return render_template('register.html', form=form)


@app.route('/login', methods=['GET', 'POST'])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        global online_user
        db_sess = db_session.create_session()
        user = db_sess.query(User).filter(User.email == form.email.data).first()
        online_user = user

        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember_me.data)
            return redirect("/home")
        return render_template('login.html',
                               message="Неправильный логин или пароль",
                               form=form)
    return render_template('login.html', form=form)


@app.route('/profile')
def profile():
    return render_template('profile.html', name=online_user.name, classes=online_user.classes,
                           school=online_user.school, type=online_user.types)


file_path = 'grades.xlsx'


def get_lessosns(date):
    conn = sqlite3.connect('db/schedule.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM lessons WHERE school = ? AND class = ? AND day = ?',
                   (online_user.school, online_user.classes, date))
    lesson = cursor.fetchall()
    conn.close()
    lesson = sorted(lesson, key=lambda x: x[2])
    return lesson


@app.route('/')
def g():
    return render_template('dnevnik.html')


@app.route('/teachers')
def teachers():
    conn = sqlite3.connect('db/teachers.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM teacher')
    t = cursor.fetchall()
    print(t)
    conn.commit()
    conn.close()
    return render_template('teachers.html', teachers=t)


@app.route('/add_teacher', methods=['POST'])
def add_teachers():
    teacher = request.form['teacher']
    subject = request.form['subject']
    school = request.form['school']

    conn = sqlite3.connect('db/teachers.db')
    cursor = conn.cursor()
    cursor.execute(
        'INSERT INTO teacher(name, school, subject) VALUES (?, ?, ?)',
        (teacher, school, subject))
    conn.commit()
    conn.close()

    return redirect('/teachers')


@app.route('/delete_teacher/<int:teacher_id>', methods=['POST'])
def delete_teacher(teacher_id):
    conn = sqlite3.connect('db/teachers.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM lessons WHERE id = ?', (teacher_id,))
    conn.commit()
    conn.close()

    return redirect('/teachers')


@app.route('/home', methods=['GET', 'POST'])
def index():
    try:
        if online_user:
            db_sess = db_session.create_session()
            user = db_sess.query(User).filter(User.email == online_user.email).first()
            if online_user.types == 'админ':
                conn = sqlite3.connect('db/schedule.db')
                cursor = conn.cursor()
                cursor.execute('SELECT * FROM lessons')
                lessons = cursor.fetchall()
                conn.close()
                return render_template('index.html', lessons=lessons)
            elif online_user.types == 'Учитель':
                global selected_value, r, e
                if e == 0:
                    m()
                    e = 1
                if selected_value != r:
                    conn = sqlite3.connect('db/users.db')
                    cursor = conn.cursor()
                    cursor.execute('SELECT * FROM users WHERE school=? AND classes = ? and name NOT LIKE ?',
                                   (online_user.school, selected_value, online_user.name))
                    studentt = cursor.fetchall()
                    r = selected_value
                    for el in studentt:
                        students[el[1]] = dict()
                        for ele in urok:
                            students[el[1]][ele] = []
                return render_template('grades.html', students=students, class_list=clas, less_list=urok)
            else:
                file_path = 'grades.xlsx'  # Путь к файлу Excel с оценками и датами
                data = get_grades(file_path, datetime.now().strftime("%Y-%m-%d"))
                lessos = get_lessosns(get_day_of_week(datetime.now().strftime("%Y-%m-%d")))
                if request.method == 'POST':
                    date = request.form['date']
                    # Здесь нужно будет получить расписание на выбранную дату (например из базы данных)
                    # и передать его в шаблон для отображения
                    # schedule = get_schedule(date)
                    lessos = get_lessosns(get_day_of_week(date))
                    data = get_grades(file_path, date.replace('.', '-'))
                print(data)
                return render_template('home.html', lessons=lessos, grade_list=data)
    except NameError:
        return redirect('/')


@app.route('/class', methods=['POST'])
def c():
    global selected_value
    selected_value = request.form['clas']
    return redirect('/home')


@app.route('/submit', methods=['POST'])
def submit():
    global selected_value
    selected_value = request.form['clas']
    return redirect('/home')


@app.route('/grade', methods=['POST'])
def grade():
    global selected_value
    student = request.form['student']
    subject = request.form['subject']
    grade = int(request.form['grade'])
    time = request.form['time']

    if student in students and subject in students[student]:
        students[student][subject].append(str(grade))
        date = datetime.now().strftime("%Y-%m-%d")
        ws.append([student, subject, grade, time, date])
        wb.save('grades.xlsx')

    return redirect('/home')


@app.route('/add_lesson', methods=['POST'])
def add_lesson():
    day = request.form['day'].capitalize()
    number = request.form['number']
    subject = request.form['subject']
    time = request.form['time']
    teacher = request.form['ticher']
    school = request.form['school']
    classs = request.form['class']

    conn = sqlite3.connect('db/schedule.db')
    cursor = conn.cursor()
    cursor.execute(
        'INSERT INTO lessons (day,number, subject, time, ticher, school, class) VALUES (?, ?, ?, ?, ?, ?, ?)',
        (day, number, subject, time, teacher, school, classs))
    conn.commit()
    conn.close()

    return redirect('/home')


def get_grades(file_path, date):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = {}

    for row in sheet.iter_rows(values_only=True):
        name, subject, grade, dat = row[:4]  # Предполагается, что дата и оценка идут первыми двумя столбцами
        dat = str(dat).split()[0]
        print(dat, date, dat == str(date))
        if dat == str(date) and name == online_user.name:
            if subject not in data:
                data[subject] = []
            data[subject].append(str(grade))

    return data


@app.route('/delete_lesson/<int:lesson_id>', methods=['POST'])
def delete_lesson(lesson_id):
    conn = sqlite3.connect('db/schedule.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM lessons WHERE id = ?', (lesson_id,))
    conn.commit()
    conn.close()

    return redirect('/home')


if __name__ == '__main__':
    db_session.global_init("db/users.db")
    app.run(debug=True)
